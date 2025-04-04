from cltk import NLP
from cltk.alphabet.lat import remove_macrons, JVReplacer
from openpyxl import load_workbook


def normalize_text(text):
	replacer = JVReplacer()
	text = remove_macrons(text)
	text = replacer.replace(text)
	return text.lower()

def deduplicate(items):
	seen = set()
	for item in items:
		if item.lemma not in seen:
			seen.add(item.lemma)
			yield item

def compare_lists(wordlist1, wordlist2):
	sharedWords = [w for w in wordlist1 if w in wordlist2]
	return sharedWords


def combine_lists(wordlist1, wordlist2):
	return set(wordlist1 + wordlist2)


def check_coverage(text, wordlist):
	total_len = len(text)


def check_vocab(filename, Vocab1, Vocab2, corelist):
	file = open(filename, 'r', encoding='utf-8')
	text = file.read()
	cltk = NLP(language="lat", suppress_banner=True)
	doc = (cltk.analyze(normalize_text(text)))
	word_objects = [word for word in doc.words if word.upos != "PUNCT"]  # list of all word objects in text
	unique_words = list(deduplicate(word_objects))
	result = []
	for word in unique_words:
		if word.lemma not in Vocab1 and word.lemma not in Vocab2:
			result.append(word)
	print("Analysing ", filename)
	for word in result:
		print(word.string, word.features, word.lemma, word.pos)

	important = [word.lemma for word in result if word.lemma in corelist]
	print("Vocab to be learned: ", important)
	return important


with open("../data/10vocab.txt", 'r', encoding="utf-8") as file:
	input_data = file.readlines()
	raw_Vocab10 = [x.split()[0] for x in input_data]
	Vocab10 = [normalize_text(i) for i in raw_Vocab10]

wb = load_workbook("../data/CLC Vocab Pool.xlsx")
ws = wb.active
raw_clc_list = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row)]
clc_list = []
try:
	for i in raw_clc_list:
		clc_list.append(normalize_text(i))
except AttributeError:
	print(i)

wb = load_workbook("../data/Latin Core Vocab.xlsx")
ws = wb.active
raw_dcc_list = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row)]
dcc_list = [normalize_text(i) for i in raw_dcc_list]
