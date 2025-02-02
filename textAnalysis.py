import multiprocessing
import os
from multiprocessing import Process
from cltk import NLP
from cltk.alphabet.lat import remove_macrons, JVReplacer
import mysql.connector as SQL
from openpyxl.reader.excel import load_workbook
import time
from semantic_text_splitter import TextSplitter
import pickle

# TODO grab random unseen passages from corpora and assess coverage
# TODO add stage selection for CLC
# TODO analyse the top words coming up that are still unknown on DCC and what words in DCC are not coming up. (Per author?)
# TODO find passages from corpora that meet a certain coverage threshold
# TODO how many words do you typically need before a word repeats in latin?
# TODO create normalised lists vocab for major textbooks
# TODO create frequency list for 90% coverage of common HS texts only, compare to textbook lists


def getText(title):
	'''
	Get full text from DB
	:param string title: Title of work
	:return: string of text
	'''
	title = title.lower()
	textQuery = "SELECT textString, authorName FROM texts WHERE LOWER(title) = %s"
	vals = (title,)
	mycursor.execute(textQuery, vals)
	results = mycursor.fetchall()
	if len(results) < 1:
		print("No text by that name found in the database.")
		return False
	if len(results) > 1:
		print("There are {} texts by that name.".format(len(results)))
		for i in range(len(results)):
			print("{}. {} {}".format(i, results[i][1], title))
		selection = int(input("Please type the number of the text you want."))
		text = results[selection][0]
		author = results[selection][1]
	# TODO offer suggestions of nearby titles if no match
	else:
		text = results[0][0]
		author = results[0][1]
	print("Text: {} by {} successfully collected".format(title, author))
	return text, author


def splitText(text, num):
	splitter = TextSplitter(len(text) // num - 1)
	chunks = splitter.chunks(text)
	return chunks


def analyse(text, return_list=None):
	cltk = NLP(language="lat", suppress_banner=True)
	print("Beginning analysis. Please wait.")
	doc = cltk.analyze(text)
	print("Finished analysing")
	if return_list:
		return_list.append(doc)
	else:
		return doc


def store_data(title, author, nlp_doc):
	if not os.path.isdir(f'docs/{author}'):
		os.mkdir(f'docs/{author}')
	with open(f'docs/{author}/{title}.pickle', 'wb') as handle:
		pickle.dump(nlp_doc, handle, protocol=pickle.HIGHEST_PROTOCOL)


def normalize_text(text):
	replacer = JVReplacer()
	text = remove_macrons(text)
	text = replacer.replace(text)
	replace_strings = ["- ", "-", "†"]
	for r in replace_strings:
		text = text.replace(r, "")
	# TODO still not catching all hyphenated words
	return text.lower()


def ignore(wordObject):
	'''
	Ignore words which are punctuation, empty strings, proper nouns, or (English) numbers
	:param wordObject:
	:return:
	'''
	if wordObject.upos == "PUNCT" or wordObject.upos == "X" or wordObject.upos == "PROPN":
		return True
	if wordObject.string == "":
		return True
	for char in wordObject.string:
		if char in "0123456789":
			return True
	return False


def deduplicate(items):
	seen = set()
	for item in items:
		if item.lemma not in seen:
			seen.add(item.lemma)
			yield item


def isFeatureInstance(wordObject, feature, value):
	keys = [str(key) for key in wordObject.features.keys()]
	if feature in keys:
		featureValue = str(wordObject.features[feature][0])
		if featureValue == value or value in featureValue:
			return True
		else:
			return False


def compare_lists(wordlist1, wordlist2):
	sharedWords = [w for w in wordlist1 if w in wordlist2]
	return sharedWords


def combine_lists(wordlist1, wordlist2):
	return set(wordlist1 + wordlist2)


def check_coverage(word_objects, vocablist):
	totalWords = len(word_objects)
	total_sharedWords = sum(1 for w in word_objects if w.lemma in vocablist)
	if total_sharedWords < 1:
		word_coverage = 0
	else:
		word_coverage = (total_sharedWords / totalWords) * 100

	lemmaList = set([w.lemma for w in word_objects])
	totalLemmata = len(lemmaList)
	sharedLemmata = compare_lists(lemmaList, vocablist)
	lemmaCoverage = (len(sharedLemmata) / totalLemmata) * 100

	unknown_words = [w.string for w in word_objects if w.lemma not in vocablist]

	return word_coverage, lemmaCoverage, unknown_words


def set_lists(wordList):
	cases = {}
	pos = {}
	verbforms = {}

	# TODO not finding any gerundives, gerunds, supines
	verbforms["subjunctives"] = [(w.string, w.index_char_start) for w in wordList if
								 isFeatureInstance(w, "Mood", "subjunctive")]
	verbforms["imperatives"] = [(w.string, w.index_char_start) for w in wordList if
								isFeatureInstance(w, "Mood", "imperative")]
	verbforms["gerundives"] = [(w.string, w.index_char_start) for w in wordList if
							   isFeatureInstance(w, "VerbForm", "gerundive")]
	verbforms["gerunds"] = [(w.string, w.index_char_start) for w in wordList if
							isFeatureInstance(w, "VerbForm", "gerund")]
	verbforms["supines"] = [(w.string, w.index_char_start) for w in wordList if
							isFeatureInstance(w, "VerbForm", "supine")]
	verbforms["participles"] = [(w.string, w.index_char_start) for w in wordList if
								isFeatureInstance(w, "VerbForm", "participle")]
	verbforms["AbAbs"] = [(w.string, w.index_char_start) for w in wordList if
						  isFeatureInstance(w, "VerbForm", "participle") and isFeatureInstance(w, "Case", "ablative")]
	verbforms["infinitives"] = [(w.string, w.index_char_start) for w in wordList if
								isFeatureInstance(w, "VerbForm", "infinitive")]
	verbforms["finite verbs"] = [(w.string, w.index_char_start) for w in wordList if
								 isFeatureInstance(w, "VerbForm", "finite")]

	# TODO not finding vocatives
	cases["ablatives"] = [(w.string, w.index_char_start) for w in wordList if isFeatureInstance(w, "Case", "ablative")]
	cases["datives"] = [(w.string, w.index_char_start) for w in wordList if isFeatureInstance(w, "Case", "dative")]
	cases["genitives"] = [(w.string, w.index_char_start) for w in wordList if isFeatureInstance(w, "Case", "genitive")]
	cases["locatives"] = [(w.string, w.index_char_start) for w in wordList if isFeatureInstance(w, "Case", "locative")]
	cases["vocatives"] = [(w.string, w.index_char_start) for w in wordList if isFeatureInstance(w, "Case", "vocative")]

	# TODO not finding proper nouns, interjections
	pos["verbs"] = [w for w in wordList if str(w.pos) == "verb"]
	pos["verbs"] += [w for w in wordList if str(w.pos) == "auxiliary"]
	pos["nouns"] = [w for w in wordList if str(w.pos) == "noun"]
	pos["proper nouns"] = [w for w in wordList if str(w.pos) == "proper_noun"]
	pos["adjectives"] = [w for w in wordList if str(w.pos) == "adjective"]
	pos["adjectives"] += [w for w in wordList if str(w.pos) == "numeral"]
	pos["pronouns"] = [w for w in wordList if str(w.pos) == "pronoun"]
	pos["prepositions"] = [w for w in wordList if str(w.pos) == "adposition"]
	pos["adverbs"] = [w for w in wordList if str(w.pos) == "adverb"]
	pos["conjunctions"] = [w for w in wordList if
						   str(w.pos) == "subordinating_conjunction" or str(w.pos) == "coordinating_conjunction"]
	pos["particles"] = [w for w in wordList if str(w.pos) == "particle"]
	pos["determiners"] = [w for w in wordList if str(w.pos) == "determiner"]
	pos["interjections"] = [w for w in wordList if str(w.pos) == "interjection"]

	return verbforms, cases, pos


def set_wordlist(name, listSource):
	"""

	:param listSource: either a filepath to a spreadsheet
	 					or a list object
	Latin terms must be in first column of spreadsheet
	:return: normalised version of list
	"""
	if os.path.exists(listSource):
		wb = load_workbook(listSource)
		ws = wb.active
		raw_list = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row)]
		result = [normalize_text(i) for i in raw_list]
	elif type(listSource) == list:
		result = [normalize_text(i) for i in listSource]
	filename = f'data/wordlists/{name} list.txt'
	with open(filename, "a") as f:
		for item in result:
			f.write(item)
		f.close()
	return result


if __name__ == "__main__":
	print("Starting...\n")

	# connect to database
	mydb = SQL.connect(
		host="localhost",
		user="root",
		password="admin",
		database="corpus"
	)
	mycursor = mydb.cursor()
	print("Connected to database.\n")

	# TODO alternate entry points - user input or hard-coded
	# get title from user, get text from db
	while True:
		title = input("Enter title of text: ")
		text, author = getText(title)
		if not text:
			title = input("Enter title of text: ")
		else:
			break

	# Split text if long then analyse each chunk
	# Store analysed chunks in shared variable
	length = len(text.split())
	num = length//1500
	if num > 2:
		chunks = splitText(text, 5)
		start_time = time.time()
		manager = multiprocessing.Manager()
		docs = manager.list()
		processes = [Process(target=analyse, args=(ch, docs)) for ch in chunks]
		for process in processes:
			process.start()
		for process in processes:
			process.join()
		print("Analysis took {} minutes.".format(round((time.time() - start_time) / 60), 2))
	else:
		start_time = time.time()
		analyse(text)
		print("Analysis took {} minutes.".format(round((time.time() - start_time) / 60), 2))

	# pickle & store
	# TODO save in database
	for i in range(len(docs)):
		store_data(title + str(i), author, docs[i])
		# TODO didn't store pickle files, but no error

	# set vocab lists
	dcc_list = set_wordlist("dcc", "data/Latin Core Vocab.xlsx")
	clc_list = set_wordlist("clc", "data/CLC Vocab Pool.xlsx")
	#
	# # get lists of text features
	# words = [word for doc in docs for word in doc.words if not ignore(word)]
	# lemmalist = set([w.lemma for w in words])
	# verbforms, cases, pos = set_lists(words)
	#
	# # print analysis
	# print("\n Totals")
	# print("Total words in text: {}".format(len(words)))
	# print("Total number of lemmata: {}".format(len(lemmalist)))
	# wordCoverage, lemmaCoverage, unknown = check_coverage(words, dcc_list)
	# print("\n Coverage")
	# print("Percentage of lemmata known: {:.2f}%".format(lemmaCoverage))
	# print("Percentage of words known: {:.2f}%".format(wordCoverage))
	# print("\n")
	#
	# print("Verb Forms")
	# for k in verbforms.keys():
	# 	print("Number of {} in text: {}".format(k, len(verbforms[k])))
	# print("\n")
	#
	# print("Noun Cases")
	# for k in cases.keys():
	# 	print("Number of {} in text: {}".format(k, len(cases[k])))
	# print("\n")
	#
	# print("Parts of Speech")
	# for k in pos.keys():
	# 	print("Number of {} in text: {}".format(k, len(pos[k])))
