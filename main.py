from cltk import NLP
from cltk.alphabet.lat import remove_macrons, JVReplacer
from openpyxl import load_workbook

# TODO grab random unseen passages from corpora and assess coverage
# TODO allow ability to paste any list of words
# TODO add stage selection for CLC
# TODO analyse the top words coming up that are still unknown on DCC and what words in DCC are not coming up. (Per author?)
# TODO find passages from corpora that meet a certain coverage threshold
# TODO how many words do you typically need before a word repeats in latin?

def read_text(textfile):
    with open(text_file, "r") as file:
        return file.read().split(" ")


def normalize_text(text):
    replacer = JVReplacer()
    text = remove_macrons(text)
    text = replacer.replace(text)
    return text.lower()


def read_wordlist(file, col):
    # Read Excel file into array
    wb = load_workbook(file)
    ws = wb.active
    return [ws.cell(row=i, column=col).value for i in range(2, ws.max_row)]


def analyze_text(text):
    cltk_nlp = NLP(language="lat", suppress_banner=True)
    return cltk_nlp.analyze(text=text)


if __name__ == '__main__':
    wordlist_file = input("Enter the filepath to your vocab list excel file:  ")
    col = int(input("Which column are your Latin words in? Enter a numeral:  "))
    raw_vocab_list = read_wordlist(wordlist_file, col)
    vocab_list = [normalize_text(i) for i in raw_vocab_list]

    punctuation = " .,;:'?! ()[]_-"

    text_file = input("Enter text filepath:  ")
    text_words = read_text(text_file) #returns a list of words
    text_length = len([x for x in text_words if x not in punctuation])    # total words
    print("text length = {}".format(text_length))
    text = " ".join(text_words)
    doc = (analyze_text(normalize_text(text)))

    word_objects = [word for word in doc.words if word.lemma not in punctuation]
    unknown_words_total = sum(1 for word in word_objects if word.lemma not in vocab_list)
    coverage = ((text_length - unknown_words_total) / text_length) * 100
    lemmata = set([word.lemma for word in word_objects])
    unknown_words = []
    for word in word_objects:
        if word.lemma not in vocab_list and word not in unknown_words:
            unknown_words.append(word)

    print("Coverage = {}%".format(coverage))
    print("Number of unknown words: {}".format(len(unknown_words)))
    unknown_words.sort(key=lambda x: x.lemma)
    for word in unknown_words:
        print(word.lemma, word.pos)
