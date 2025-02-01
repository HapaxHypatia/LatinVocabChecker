from cltk import NLP
from cltk.alphabet.lat import remove_macrons, JVReplacer
from openpyxl import load_workbook
import os.path


def read_text(text_input):
    '''
    :return: list of strings
    '''
    if os.path.exists(text_input):
        with open(text_input, "r") as file:
            return file.read()
    elif type(text_input) == str:
        return text_input


def normalize_text(text):
    replacer = JVReplacer()
    text = remove_macrons(text)
    text = replacer.replace(text)
    return text.lower()


def read_wordlist(list_input):
    '''
    :return: list of strings
    '''
    if os.path.exists(list_input):
        col = int(input("Which column are your Latin words in? Enter a numeral:  "))
        wb = load_workbook(list_input)
        ws = wb.active
        raw_list = [ws.cell(row=i, column=col).value for i in range(2, ws.max_row)]
    elif type(list_input) == str:
        raw_list = list_input.split(',')
    normalized_vocab_list = [normalize_text(i) for i in raw_list]
    vocab_objects = cltk.analyze(" ".join(normalized_vocab_list))
    return [word.lemma for word in vocab_objects]
    return [word.lemma for word in normalized_vocab_list]

if __name__ == '__main__':
    cltk = NLP(language="lat", suppress_banner=True)
    punctuation = " .,;:'\"?! ()[]_-"

    text_input = input("Enter text filepath or text:  ")
    text = read_text(text_input)  # returns a list of words (strings)

    list_input = input("Enter a word list separated by commas, or the filepath to your vocab list excel file:  ")
    vocab_list = read_wordlist(list_input)

    doc = (cltk.analyze(normalize_text(text)))
    word_objects = [word for word in doc.words if word.upos != "PUNCT"]  # list of all word objects in text

    text_length = len(word_objects)
    print("text length = {}".format(text_length))
    total_unknown_tokens = sum(1 for word in word_objects if word.lemma not in vocab_list)  # int

    lemmata = sorted(set([lemma for lemma in doc.lemmata]))  # set of strings: all lemmata in text
    total_lemmata = len(lemmata)  # int
    unknown_lemmata = []
    for word in word_objects:
        if word.lemma in vocab_list or word in unknown_lemmata:
            continue
        unknown_lemmata.append(word)
    unknown_lemmata.sort(key=lambda x: x.lemma)
    total_unknown_lemmata = len(unknown_lemmata)

    lemma_coverage = ((total_lemmata - total_unknown_lemmata) / total_lemmata) * 100
    token_coverage = ((text_length - total_unknown_tokens) / text_length) * 100

    # TODO: Grammatical analysis. List of grammar features in text (Tenses etc)

    print("Token Coverage = {}%".format(token_coverage))
    print("Lemma Coverage = {}%".format(lemma_coverage))
    print("Number of unknown lemmata: {}".format(total_unknown_lemmata))
    for item in unknown_lemmata:
        print(item.lemma, item.pos)
    print("Vocab to be learned: ")
    for item in unknown_lemmata:
        if item in read_wordlist("data/Latin Core Vocab.xlsx"):
            print(item)
    #TODO including duplicate lemmata
