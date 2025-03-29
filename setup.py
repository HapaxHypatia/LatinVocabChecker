import multiprocessing
import os
from multiprocessing import Process
from openpyxl.reader.excel import load_workbook
import time
from semantic_text_splitter import TextSplitter
import pickle
import re
from unidecode import unidecode
from cltk.data.fetch import FetchCorpus
from cltk import NLP
from DB import getText
import re


def getCorpora():
	corpus_downloader = FetchCorpus(language="lat")
	corpora = corpus_downloader.list_corpora
	for c in corpora:
		try:
			corpus_downloader.import_corpus(c)
			print(f"{c} successfully downloaded.")
		except Exception as e:
			print(c)
			print(e)
			continue


def splitText(text, num):
	print(f"Splitting Text into {num} parts")
	splitter = TextSplitter(len(text) // num)
	chunks = splitter.chunks(text)
	result = [(ind, ch) for ind, ch in enumerate(chunks)]
	print(f"Returning {len(chunks)} chunks.")
	return result


def analyseLarge(chunk, return_list):
	ind, text = chunk
	cltk = NLP(language="lat", suppress_banner=True)
	print(f"Beginning analysis of chunk {ind} of large text. Please wait.")
	doc = cltk.analyze(text)
	print(f"Finished analysing chunk {ind} of large text")
	return_list.append((ind, doc))


def analyseSmall(text):
	cltk = NLP(language="lat", suppress_banner=True)
	print("Beginning analysis of small text. Please wait.")
	doc = cltk.analyze(text)
	print("Finished analysing")
	return doc


def store_data(author, title, part, nlp_doc):
	"""

	:param author: string
	:param title: string
	:param part: string of a digit
	:param nlp_doc: nlp-document
	:return:
	"""
	author = author.lower()
	title = title.lower()
	if not os.path.isdir(f'docs/{author}'):
		os.mkdir(f'docs/{author}')

	if not os.path.isdir(f'docs/{author}/{title}'):
		os.mkdir(f'docs/{author}/{title}')

	with open(f'docs/{author}/{title}/{title}{part}.pickle', 'wb') as handle:
		pickle.dump(nlp_doc, handle, protocol=pickle.HIGHEST_PROTOCOL)
		print("{} by {} successfully saved".format(title, author))


def normalize_text(text):
	"""
	Removes macrons
	Replaces Vs and Js
	Removes punctuation in the middle of words
	:param text: string of a passage
	:return: string
	"""
	if type(text) != str:
		text = str(text)
	text = text.lower()

	# replace macrons
	macrons = {
		'ā': 'a',
		'ē': 'e',
		'ō': 'o',
		'ī': 'I',
		'ū': 'u',
		'ȳ': 'y'
	}
	for k, v in macrons.items():
		text.replace(k, v)

	# replace Js and Vs
	jv = {
		'ja': 'ia',
		'je': 'ie',
		'ji': 'ii',
		'jo': 'io',
		'ju': 'iu',
		'va': 'ua',
		've': 'ue',
		'vi': 'ui',
		'vo': 'uo',
		'vu': 'uu',
	}
	for k, v in jv.items():
		text.replace(k, v)

	# join words that split over line break
	text = re.sub(r'(-\s{0,3})', "", text)
	return text

def set_wordlist(name, listSource):
	"""
	:param name: string.
	:param listSource: either a filepath to a spreadsheet	 					or a list object
	Latin terms must be in first column of spreadsheet
	:return: normalised version of list
	"""
	if os.path.exists(listSource):
		wb = load_workbook(listSource)
		ws = wb.active
		raw_list = list([ws.cell(row=i, column=1).value for i in range(2, ws.max_row)])
		result = [normalize_text(i) for i in raw_list]
	elif type(listSource) == list:
		result = [normalize_text(i) for i in listSource]
	filename = f'data/wordlists/{name} list.txt'
	with open(filename, "a", encoding="utf-8") as f:
		for item in result:
			if not item:
				continue
			if len(item.split()) > 1:
				item = unidecode(item.split()[0])
			else:
				item = unidecode(item)
			f.write(item)
			f.write("\n")
		f.close()
	return result


def pickleExists(title, author):
	if os.path.isdir(f"docs/{author}"):
		for folder in os.listdir(f"docs/{author}"):
			for filename in os.listdir(f"docs/{author}/{folder}"):
				if re.search(rf"{title.lower()}\d.pickle", filename):
					return True
	else:
		return False


def getVocab(name):
	with open(f"data/wordlists/{name} list.txt", 'r', encoding='utf-8') as f:
		vocab = [x.strip() for x in f.readlines()]
		return vocab


if __name__ == "__main__":
	print("Starting...\n")

	# getCorpora()

	# set vocab lists
	vocabLists = [
		("dcc", "data/vocab/Latin Core Vocab.xlsx"),
		("clc", "data/vocab/cambridge_latin_course.xlsx"),  # Missing some Bk5 vocab
		("llpsi", "data/vocab/lingua_latina.xlsx"),
		("olc", "data/vocab/oxford_latin_course.xlsx"),  # Only to Ch.22
		("ecrom", "data/vocab/ecce_romani.xlsx"),
		("sub", "data/vocab/suburani.xlsx"),
		("wheel", "data/vocab/wheelock.xlsx"),
		("newmil", "data/vocab/latin_for_the_new_millennium.xlsx")
	]

	# get senior texts
	works = [
		("cicero", "In Catilinam"),
		("cicero", "In Pisonem"),
		("cicero", "In Q. Caecilium"),
		("cicero", "In Sallustium [sp.]"),
		("cicero", "In Vatinium"),
		("cicero", "In Verrem"),
		("cicero", "Pro Archia"),
		("cicero", "Pro Balbo"),
		("cicero", "Pro Caecina"),
		("cicero", "Pro Caelio"),
		("cicero", "Pro Cluentio"),
		("cicero", "Pro Flacco"),
		("cicero", "Pro Fonteio"),
		("cicero", "Pro Lege Manilia"),
		("cicero", "Pro Ligario"),
		("cicero", "Pro Marcello"),
		("cicero", "Pro Milone"),
		("cicero", "Pro Murena"),
		("cicero", "Pro Plancio"),
		("cicero", "Pro Q. Roscio Comoedo"),
		("cicero", "Pro Quinctio"),
		("cicero", "Pro Rabirio Perduellionis Reo"),
		("cicero", "Pro Rabirio Postumo"),
		("cicero", "Pro Rege Deiotaro"),
		("cicero", "Pro S. Roscio Amerino"),
		("cicero", "Pro Scauro"),
		("cicero", "Pro Sestio"),
		("cicero", "Pro Sulla"),
		("cicero", "Pro Tullio"),
		("Caesar", "de bello gallico"),
		("catullus", "carmina"),
		("livius", "ab urbe condita"),
		("ovidius", "metamorphoses"),
		("ovidius", "amores"),
		("ovidius", "remedia amoris"),
		("ovidius", "Epistulae (vel Heroides)"),
		("plinius", "epistulae"),
		("virgilius", "aeneis")
	]

	for title in works:
		text, author = getText(title[1], title[0])

		if pickleExists(title[1], author):
			print("Text analysis already stored.")
			continue

		# Split text if long then analyse each chunk
		length = len(text.split())
		print("Text length: {}".format(length))
		num = length//2000
		manager = multiprocessing.Manager()
		docs = manager.list()
		if num > 1:
			chunks = splitText(text, min(10, num))
			start_time = time.time()
			# NOTE: multiprocessing must be done in __main__
			processes = [Process(target=analyseLarge, args=(ch, docs)) for ch in chunks]
			for ind, process in enumerate(processes):
				process.start()
			for ind, process in enumerate(processes):
				process.join()
			print("Analysis took {} minutes.".format(round((time.time() - start_time) / 60), 2))
		else:
			start_time = time.time()
			docs.append((0, analyseSmall(text)))
			print("Analysis took {} minutes.".format(round((time.time() - start_time) / 60), 2))

		# pickle & store
		# TODO save in database
		res = sorted(docs, key=lambda x: x[0])
		for d in res:
			store_data(author, title[1], str(d[0]), d[1])

	# for k, v in vocabLists.items():
	# 	set_wordlist(v[0], v[1])

