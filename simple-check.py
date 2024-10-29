from cltk import NLP
from cltk.alphabet.lat import remove_macrons, JVReplacer
from openpyxl import load_workbook


def normalize_text(text):
	replacer = JVReplacer()
	text = remove_macrons(text)
	text = replacer.replace(text)
	return text.lower()


def check_vocab(filename, Vocab1, Vocab2, corelist):
	file = open(filename, 'r', encoding='utf-8')
	text = file.read()
	cltk = NLP(language="lat", suppress_banner=True)
	doc = (cltk.analyze(normalize_text(text)))
	word_objects = [word for word in doc.words if word.pos != 'punctuation']  # list of all word objects in text
	unique_lemmata = []
	words = []
	for word in word_objects:
		if word.lemma not in Vocab1 and word.lemma not in Vocab2:
			if word.lemma not in unique_lemmata:
				unique_lemmata.append(word.lemma)
				words.append(word)
	print("Analysing ", filename)
	for word in words:
		print(word.string, word.features, word.lemma, word.pos)

	important = [word.lemma for word in words if word.lemma in corelist]
	print("Vocab to be learned: ", important)
	return important


with open("10vocab.txt", 'r', encoding="utf-8") as file:
	input_data = file.readlines()
	raw_Vocab10 = [x.split()[0] for x in input_data]
	Vocab10 = [normalize_text(i) for i in raw_Vocab10]

wb = load_workbook("data/CLC Vocab Pool.xlsx")
ws = wb.active
raw_clc_list = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row)]
clc_list = []
try:
	for i in raw_clc_list:
		clc_list.append(normalize_text(i))
except AttributeError:
	print(i)

wb = load_workbook("data/Latin Core Vocab.xlsx")
ws = wb.active
raw_dcc_list = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row)]
dcc_list = [normalize_text(i) for i in raw_dcc_list]

Vocab10 += check_vocab("data/Thras1.txt", Vocab10, clc_list, dcc_list)
Vocab10 += check_vocab("data/Thras2.txt", Vocab10, clc_list, dcc_list)
Vocab10 += check_vocab("data/Thras3.txt", Vocab10, clc_list, dcc_list)
Vocab10 += check_vocab("data/Thras4.txt", Vocab10, clc_list, dcc_list)
